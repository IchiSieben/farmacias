"""recon/mapa_categorias.py — Mapa de categorías/subcategorías por cadena -> Excel.

Captura el ÁRBOL DE ORGANIZACIÓN de cada web, tal como cada una lo segmenta. Es
para orientar navegación/organización del comparador, NO para el matching (el
matcher empareja por principio activo + specs, no por categoría).

Fuentes (cada web organiza distinto):
  • Inkafarma / Mifarma (InRetail · Algolia): facetas `category` -> `subCategory`
    con conteos. Se usa batch multi-query (varias categorías por POST) para que
    el árbol completo cueste pocas llamadas HTTP.
  • Boticas Perú (Salesforce Commerce Cloud): navegación del sitio. El megamenú
    expone un path /nivel1/nivel2[/nivel3]; el conteo sale del `.result-count`
    ("N Resultados") de cada página de categoría.

Salida: categorias_farmacias.xlsx
  - Una hoja por cadena (Categoría | Subcategoría | Nº prod. subcat | Nº prod. cat)
  - Hoja "Resumen": nº de categorías, subcategorías y tamaño de catálogo por web.

Uso:
    py -m recon.mapa_categorias
    py -m recon.mapa_categorias --salida categorias_farmacias.xlsx

Python 3.9+. Dependencias: httpx, selectolax, openpyxl, pyyaml.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.parse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from selectolax.parser import HTMLParser

from core.adapters.boticasperu import BoticasPeruAdapter
from core.adapters.inkafarma import InkafarmaAdapter
from core.adapters.mifarma import MifarmaAdapter

ROOT = Path(__file__).resolve().parents[1]
OUT_DEFAULT = ROOT / "categorias_farmacias.xlsx"

NOMBRE_CADENA = {
    "inkafarma": "Inkafarma",
    "mifarma": "Mifarma",
    "boticasperu": "Boticas Perú",
}


# ============================ InRetail (Algolia) ============================
def _facet_categorias(adapter, solo_web: bool = True) -> Tuple[Dict[str, int], int]:
    """Faceta `category` global: {categoria: nº productos} + total del catálogo."""
    url = f"{adapter.host}/1/indexes/*/queries"
    extra = "&facets=" + urllib.parse.quote(json.dumps(["category"])) + "&maxValuesPerFacet=1000"
    body = {"requests": [{
        "indexName": adapter.index,
        "params": adapter._params("", 0, 0, adapter._filtros(solo_web), extra=extra),
    }]}
    res = adapter._post_json(url, body)["results"][0]
    cats = res.get("facets", {}).get("category", {})
    total = res.get("nbHits", 0)
    return cats, total


def _facet_subcats_batch(adapter, categorias: List[str], solo_web: bool = True,
                         lote: int = 40) -> Dict[str, Dict[str, int]]:
    """Para cada categoría, faceta `subCategory` (filtrando category:cat).

    Usa batch multi-query: hasta `lote` categorías por POST -> pocas llamadas.
    Devuelve {categoria: {subcategoria: nº productos}}.
    """
    url = f"{adapter.host}/1/indexes/*/queries"
    fextra = "&facets=" + urllib.parse.quote(json.dumps(["subCategory"])) + "&maxValuesPerFacet=1000"
    out: Dict[str, Dict[str, int]] = {}
    for i in range(0, len(categorias), lote):
        bloque = categorias[i:i + lote]
        reqs = [{
            "indexName": adapter.index,
            "params": adapter._params("", 0, 0,
                                      adapter._filtros(solo_web, [f"category:{c}"]),
                                      extra=fextra),
        } for c in bloque]
        res = adapter._post_json(url, {"requests": reqs})["results"]
        for c, r in zip(bloque, res):
            out[c] = r.get("facets", {}).get("subCategory", {})
        if i + lote < len(categorias):
            time.sleep(0.2)
    return out


def arbol_inretail(adapter_cls) -> dict:
    """Árbol category->subCategory con conteos para una cadena InRetail."""
    adapter = adapter_cls.from_yaml(delay_range=(0, 0))
    with adapter:
        cats, total = _facet_categorias(adapter)
        nombres = sorted(cats.keys())
        print(f"  {adapter.cadena}: {len(nombres)} categorías, "
              f"catálogo WEB ~{total}", file=sys.stderr)
        subs = _facet_subcats_batch(adapter, nombres)
    filas: List[dict] = []
    for cat in sorted(cats, key=lambda c: -cats[c]):
        sc = subs.get(cat, {})
        if sc:
            for sub, n in sorted(sc.items(), key=lambda x: -x[1]):
                filas.append({"categoria": cat, "subcategoria": sub,
                              "n_sub": n, "n_cat": cats[cat]})
        else:
            filas.append({"categoria": cat, "subcategoria": "(sin subcategoría)",
                          "n_sub": cats[cat], "n_cat": cats[cat]})
    distintas = {s for sc in subs.values() for s in sc}
    return {
        "cadena": adapter_cls.cadena,
        "filas": filas,
        "n_categorias": len(cats),
        "n_subcategorias": len(distintas),           # nombres de subcategoría distintos
        "n_pares": len(filas),                        # combinaciones categoría×subcategoría
        "total_catalogo": total,
    }


# ============================ Boticas Perú (SFCC) ===========================
_RE_SUFIJO = re.compile(r"-\d+$")
_RE_NUM = re.compile(r"[\d.,]+")


def _slug_base(slug: str) -> str:
    return _RE_SUFIJO.sub("", slug)


def _titulo(slug: str) -> str:
    return slug.replace("_", " ").replace("-", " ").strip().title()


def _result_count(adapter, path: str) -> Optional[int]:
    """Nº de productos de una página de categoría Boticas (.result-count)."""
    try:
        r = adapter._client.get(adapter.dominio + path, follow_redirects=True)
        if r.status_code != 200:
            return None
        node = HTMLParser(r.text).css_first(".result-count")
        if not node:
            return None
        m = _RE_NUM.search(node.text())
        return int(m.group(0).replace(",", "").replace(".", "")) if m else None
    except Exception:
        return None


def arbol_boticas() -> dict:
    """Árbol nivel1->nivel2 (categoría->subcategoría) de Boticas vía navegación."""
    adapter = BoticasPeruAdapter.from_yaml(delay_range=(0, 0))
    with adapter:
        home = adapter._client.get(adapter.dominio + "/")
        tree = HTMLParser(home.text)

        # nivel 1: encabezados del menú (nombre + href real, con sufijo -N).
        nivel1: Dict[str, dict] = {}        # base_slug -> {nombre, path}
        for n in tree.css("li.nav-item > a.nav-link"):
            href = n.attributes.get("href", "")
            if not href.startswith("/"):
                continue
            slug = href.strip("/").split("/")[0]
            base = _slug_base(slug)
            nombre = re.sub(r"\s+", " ", n.text()).strip()
            nivel1.setdefault(base, {"nombre": nombre, "path": href})

        # nivel 2 (y nivel 3) desde los enlaces del megamenú. SFCC duplica IDs con
        # sufijo (-1/-2): se colapsan por slug base y se descartan etiquetas "null".
        sub: Dict[str, Dict[str, dict]] = {}   # cat_base -> {nivel2_base: {nombre, path, ...}}
        for x in tree.css("a.dropdown-link"):
            href = x.attributes.get("href", "")
            if not href.startswith("/"):
                continue
            parts = [p for p in href.strip("/").split("/") if p]
            if len(parts) < 2:
                continue
            cat = _slug_base(parts[0])
            n2key = _slug_base(parts[1])
            raw = re.sub(r"\s+", " ", x.text()).strip()
            es_null = raw.lower() == "null" or not raw
            nombre = _titulo(n2key) if es_null else raw
            directo = len(parts) == 2          # enlace de nivel-2 directo (no derivado de nivel-3)
            cand = {"nombre": nombre, "path": "/" + "/".join(parts[:2]),
                    "directo": directo, "null": es_null}
            nivel1.setdefault(cat, {"nombre": _titulo(cat), "path": "/" + parts[0]})
            d = sub.setdefault(cat, {})
            cur = d.get(n2key)
            # Preferir el enlace directo de nivel-2 y el nombre real (no 'null').
            if cur is None or (directo and not cur["directo"]) or (cur["null"] and not es_null):
                d[n2key] = cand

        # conteos: una página por categoría (nivel1) y por subcategoría (nivel2).
        total_n2 = sum(len(v) for v in sub.values())
        print(f"  boticasperu: {len(nivel1)} categorías, {total_n2} subcategorías "
              f"-> consultando conteos ({len(nivel1) + total_n2} páginas)...",
              file=sys.stderr)

        cat_count: Dict[str, Optional[int]] = {}
        for i, (cat, meta) in enumerate(sorted(nivel1.items())):
            cat_count[cat] = _result_count(adapter, meta["path"])
            time.sleep(0.15)

        filas: List[dict] = []
        hechos = 0
        for cat in sorted(nivel1, key=lambda c: -(cat_count.get(c) or 0)):
            meta1 = nivel1[cat]
            subs = sub.get(cat, {})
            if not subs:
                filas.append({"categoria": meta1["nombre"],
                              "subcategoria": "(sin subcategoría)",
                              "n_sub": cat_count.get(cat), "n_cat": cat_count.get(cat)})
                continue
            conteos = []
            for n2slug, m2 in subs.items():
                n = _result_count(adapter, m2["path"])
                conteos.append((m2["nombre"], n))
                hechos += 1
                if hechos % 25 == 0:
                    print(f"    ... {hechos}/{total_n2} subcategorías", file=sys.stderr)
                time.sleep(0.15)
            for nombre2, n in sorted(conteos, key=lambda x: -(x[1] or 0)):
                filas.append({"categoria": meta1["nombre"], "subcategoria": nombre2,
                              "n_sub": n, "n_cat": cat_count.get(cat)})

    return {
        "cadena": "boticasperu",
        "filas": filas,
        "n_categorias": len(nivel1),
        "n_subcategorias": total_n2,
        "n_pares": total_n2,      # en Boticas el árbol es estricto (1 subcat -> 1 cat)
        "total_catalogo": None,   # SFCC no expone un total global limpio
    }


# ============================ Excel =========================================
def escribir_excel(arboles: List[dict], salida: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="0D2B45")
    hdr_font = Font(bold=True, color="FFFFFF")
    cat_fill = PatternFill("solid", fgColor="E7EEF5")
    bold = Font(bold=True)

    def estilo_encabezado(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    # --- Resumen ---
    ws = wb.create_sheet("Resumen")
    ws.append(["Cadena", "Nº categorías", "Subcategorías (distintas)",
               "Combinaciones cat×subcat", "Catálogo (productos)", "Modelo de árbol", "Fuente"])
    estilo_encabezado(ws, 7)
    fuente = {"inkafarma": "Algolia (facetas category/subCategory)",
              "mifarma": "Algolia (facetas category/subCategory)",
              "boticasperu": "Navegación del sitio (SFCC) + .result-count"}
    modelo = {"inkafarma": "Facetas many-to-many (1 subcat en varias cat)",
              "mifarma": "Facetas many-to-many (1 subcat en varias cat)",
              "boticasperu": "Árbol estricto por path /nivel1/nivel2"}
    for a in arboles:
        ws.append([NOMBRE_CADENA[a["cadena"]], a["n_categorias"], a["n_subcategorias"],
                   a.get("n_pares", len(a["filas"])),
                   a["total_catalogo"] if a["total_catalogo"] is not None else "n/d",
                   modelo.get(a["cadena"], ""), fuente.get(a["cadena"], "")])
    ws.append([])
    nota = ws.cell(row=ws.max_row + 1, column=1,
                   value="Mapa de ORGANIZACIÓN de cada web (cómo segmenta su catálogo). "
                         "No se usa para el matching (que empareja por principio activo). "
                         "Inka/Mifarma exponen facetas Algolia many-to-many: una misma "
                         "subcategoría cuelga de varias categorías, por eso 'combinaciones' "
                         ">> 'subcategorías distintas' y un producto puede contarse en varias. "
                         "Boticas organiza un árbol estricto por path; el conteo es 'N "
                         "Resultados' de cada página de categoría.")
    nota.font = Font(italic=True, color="666666")
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=nota.row, start_column=1, end_row=nota.row, end_column=7)
    ws.row_dimensions[nota.row].height = 75
    for col, w in zip("ABCDEFG", (16, 14, 20, 22, 20, 38, 42)):
        ws.column_dimensions[col].width = w

    # --- una hoja por cadena ---
    for a in arboles:
        ws = wb.create_sheet(NOMBRE_CADENA[a["cadena"]][:31])
        ws.append(["Categoría", "Subcategoría", "Nº prod. (subcat)", "Nº prod. (categoría)"])
        estilo_encabezado(ws, 4)
        cat_prev = None
        for f in a["filas"]:
            primera = f["categoria"] != cat_prev
            cat_prev = f["categoria"]
            ws.append([f["categoria"] if primera else "", f["subcategoria"],
                       f["n_sub"] if f["n_sub"] is not None else "",
                       f["n_cat"] if (primera and f["n_cat"] is not None) else ""])
            if primera:
                for c in range(1, 5):
                    ws.cell(row=ws.max_row, column=c).fill = cat_fill
                ws.cell(row=ws.max_row, column=1).font = bold
                ws.cell(row=ws.max_row, column=4).font = bold
        for col, w in zip("ABCD", (30, 40, 16, 18)):
            ws.column_dimensions[col].width = w

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Mapa de categorías por cadena -> Excel.")
    parser.add_argument("--salida", default=str(OUT_DEFAULT))
    parser.add_argument("--solo", default=None,
                        help="Limitar a cadenas (coma): inkafarma,mifarma,boticasperu")
    args = parser.parse_args(argv)

    quiero = set((args.solo or "inkafarma,mifarma,boticasperu").split(","))
    arboles: List[dict] = []
    print("Construyendo mapa de categorías...", file=sys.stderr)
    if "inkafarma" in quiero:
        arboles.append(arbol_inretail(InkafarmaAdapter))
    if "mifarma" in quiero:
        arboles.append(arbol_inretail(MifarmaAdapter))
    if "boticasperu" in quiero:
        arboles.append(arbol_boticas())

    out = Path(args.salida)
    escribir_excel(arboles, out)
    print(f"\nListo: {out}")
    for a in arboles:
        print(f"  {NOMBRE_CADENA[a['cadena']]:14} {a['n_categorias']:>3} cat · "
              f"{a['n_subcategorias']:>4} subcat · "
              f"catálogo {a['total_catalogo'] if a['total_catalogo'] is not None else 'n/d'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
